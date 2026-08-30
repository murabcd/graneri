from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname, range_to_tuple
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="E8EEF8")
HEADER_FONT = Font(bold=True, color="172033")


def _write_rows(worksheet: Worksheet, rows: list[list[Any]]) -> None:
    for row in rows:
        worksheet.append(row)


def _style_sheet(worksheet: Worksheet, frozen_rows: int) -> None:
    if frozen_rows > 0:
        worksheet.freeze_panes = f"A{frozen_rows + 1}"
    for cell in worksheet[1]:
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(vertical="center")
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        maximum = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=0,
        )
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(max(maximum + 2, 10), 48)


def _add_charts(worksheet: Worksheet, charts: list[dict[str, Any]]) -> None:
    header_cells = [cell for cell in worksheet[1] if cell.value is not None]
    header_names = [str(cell.value) for cell in header_cells]
    if charts and len(header_names) != len(set(header_names)):
        raise ValueError(
            f"Charts require unique header names on sheet {worksheet.title}."
        )
    headers = {str(cell.value): cell.column for cell in header_cells}
    for chart_spec in charts:
        category_column = headers.get(chart_spec["categoryColumn"])
        data_columns = [headers.get(name) for name in chart_spec["dataColumns"]]
        if category_column is None or any(column is None for column in data_columns):
            raise ValueError(
                f"Chart columns were not found on sheet {worksheet.title}."
            )
        chart_type = chart_spec["type"]
        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            raise ValueError(f"Unsupported spreadsheet chart: {chart_type}")
        categories = Reference(
            worksheet,
            min_col=category_column,
            min_row=2,
            max_row=worksheet.max_row,
        )
        for data_column in data_columns:
            data = Reference(
                worksheet,
                min_col=data_column,
                min_row=1,
                max_row=worksheet.max_row,
            )
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = chart_spec["title"]
        chart.height = 8
        chart.width = 14
        worksheet.add_chart(chart, chart_spec.get("position", "H2"))


def _populate_sheet(worksheet: Worksheet, spec: dict[str, Any]) -> None:
    _write_rows(worksheet, spec["rows"])
    _style_sheet(worksheet, spec.get("frozenRows", 1))
    _add_charts(worksheet, spec.get("charts", []))


def _extend_chart_reference(
    reference: Any, worksheet: Worksheet, old_last_row: int
) -> None:
    if reference is None or not isinstance(reference.f, str) or not reference.f:
        return
    try:
        sheet_name, (min_column, min_row, max_column, max_row) = range_to_tuple(
            reference.f
        )
    except ValueError:
        return
    if sheet_name != worksheet.title or max_row != old_last_row:
        return
    reference.f = (
        f"{quote_sheetname(sheet_name)}!"
        f"${get_column_letter(min_column)}${min_row}:"
        f"${get_column_letter(max_column)}${worksheet.max_row}"
    )
    if hasattr(reference, "numCache"):
        reference.numCache = None
    if hasattr(reference, "strCache"):
        reference.strCache = None


def _extend_charts_after_append(worksheet: Worksheet, old_last_row: int) -> None:
    for chart in worksheet._charts:  # noqa: SLF001
        for series in chart.series:
            for data_source in (
                getattr(series, "cat", None),
                getattr(series, "val", None),
                getattr(series, "xVal", None),
                getattr(series, "yVal", None),
            ):
                if data_source is None:
                    continue
                _extend_chart_reference(
                    getattr(data_source, "numRef", None),
                    worksheet,
                    old_last_row,
                )
                _extend_chart_reference(
                    getattr(data_source, "strRef", None),
                    worksheet,
                    old_last_row,
                )


def create_spreadsheet(sheets: list[dict[str, Any]], destination: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_spec in sheets:
        worksheet = workbook.create_sheet(sheet_spec["name"])
        _populate_sheet(worksheet, sheet_spec)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(destination)


def edit_spreadsheet(
    source: Path, edits: list[dict[str, Any]], destination: Path
) -> None:
    workbook = load_workbook(source)
    changed_sheets: set[str] = set()
    for edit in edits:
        edit_type = edit["type"]
        if edit_type == "add_sheet":
            sheet_spec = edit["sheet"]
            if sheet_spec["name"] in workbook.sheetnames:
                raise ValueError(f"Sheet already exists: {sheet_spec['name']}")
            worksheet = workbook.create_sheet(sheet_spec["name"])
            _populate_sheet(worksheet, sheet_spec)
            continue
        sheet_name = edit["sheet"]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Spreadsheet sheet was not found: {sheet_name}")
        worksheet = workbook[sheet_name]
        changed_sheets.add(sheet_name)
        if edit_type == "append_rows":
            old_last_row = worksheet.max_row
            _write_rows(worksheet, edit["rows"])
            _extend_charts_after_append(worksheet, old_last_row)
        elif edit_type == "set_cell":
            worksheet[edit["cell"]] = edit["value"]
        else:
            raise ValueError(f"Unsupported spreadsheet edit: {edit_type}")
    for sheet_name in changed_sheets:
        _style_sheet(workbook[sheet_name], 1)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(destination)
